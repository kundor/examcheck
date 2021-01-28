#!/usr/bin/python3

import mmap
import subprocess
from hashlib import blake2b
from contextlib import closing
from collections import Counter, defaultdict

import simhash
import IPython

from openpyxl import load_workbook
from canvas import *
from xlinfo import workbook_props
from xlsx2csv import process_cells, RowVisitor, SHEETSEP, CSVPrinter
from uniquecells import cleanval, CellCollector, filerpts, pairs_few, most_shared

def blakesum(buf):
    bsum = blake2b(buf, digest_size=24)
    return bsum.hexdigest()

def bsum_fid(fid):
    return blakesum(mmap.mmap(fid.fileno(), 0, access=mmap.ACCESS_READ))

def bsum_file(filepath):
    with filepath.open('rb') as fid:
        return bsum_fid(fid)

class BlakeHasher(RowVisitor):
    def __init__(self):
        self.bsum = blake2b(digest_size=24)
    def __call__(self, row):
        rowstr = ','.join(str(c) if c is not None else '' for c in row).rstrip(',') + '\n'
        self.bsum.update(rowstr.encode())
    def newsheet(self):
        self.bsum.update(SHEETSEP.encode())
    def value(self):
        return self.bsum.hexdigest()

class SimHasher(RowVisitor):
    def __init__(self):
        self.hashvector = []
    @staticmethod
    def hashval(shingle):
        return simhash.unsigned_hash(shingle.encode())
    def __call__(self, row):
        rowcells = [cleanval(c) for c in row if c is not None]
        if len(rowcells) >= 3:
            shingles = [' '.join(cells) for cells in simhash.shingle(rowcells, 3)]
        else:
            shingles = [' '.join(rowcells)]
        self.hashvector += [self.hashval(s) for s in shingles]
    def value(self):
        return simhash.compute(self.hashvector)

def get_mime(filepath):
    result = subprocess.run(['file', '-b', '--mime-type', str(filepath)], capture_output=True)
    return result.stdout.decode().rstrip()

def xls2xlsx(filepath):
    newpath = filepath.with_suffix('.xlsx')
    if newpath.exists():
        print(f'{newpath} found', file=sys.stderr)
    elif get_mime(filepath) == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        print("Appears to be misnamed .xlsx file", file=sys.stderr)
        filepath.rename(newpath)
    else:
        print(f'Converting {filepath} to xlsx', file=sys.stderr)
        subprocess.run(['libreoffice', '--headless',
                '--convert-to', 'xlsx',
                '--outdir', newpath.parent,
                str(filepath)],
                stdout=subprocess.DEVNULL)
    return newpath

def process_workbook(xlhash, filepath, workbook):
    outfile = filepath.with_suffix('.csv')
    cells, csvhash, simhash, _ = process_cells(workbook,
            [CellCollector(), BlakeHasher(), SimHasher(), CSVPrinter(outfile)])
    info = workbook_props(workbook, filepath.name, xlhash, csvhash, simhash)
    return cells, info

def process_file(filepath):
    with open(filepath, 'rb') as fid:
        xlhash = bsum_fid(fid)
        with closing(load_workbook(fid, read_only=True)) as workbook:
            return process_workbook(xlhash, filepath, workbook)

origfile = Path('original/Module_8_S20.xlsx').resolve()
dodir = Path('mod8')
origcells, originfo = process_file(origfile)

cellfiles = defaultdict(list) # map cell_content : files
donefiles = set()
xlhashes = {}
infos = []

warnings.filterwarnings('ignore', '.*invalid specification.*', UserWarning, 'openpyxl')
warnings.filterwarnings('ignore', 'Unknown extension is not supported.*', UserWarning, 'openpyxl')

sleeps = 0

while True:
    oldn = len(donefiles)
    for file in dodir.glob('*.xls*'):
        if file in donefiles:
            continue
        xlhash = bsum_file(file)
        if xlhash in xlhashes:
            prev = xlhashes[xlhash]
            print(f'File {file.name} identical to previously seen file {prev.filename}')
            continue
        if file.name.endswith('.xls'):
            file = xls2xlsx(file)
        try:
            wb = load_workbook(file, read_only=True)
        except Exception as e:
            print(file.name, 'is not an xlsx file?', e, file=sys.stderr)
            continue
        thecells, theinfo = process_workbook(xlhash, file, wb)
        for cval in thecells - origcells:
            cellfiles[cval].append(file.name)
        infos.append(theinfo)
        xlhashes[xlhash] = theinfo
        donefiles.add(file)
        print(file)
        wb.close()
    if len(donefiles) == oldn:
        if sleeps > 10 and yesno('All done?'):
            break
        time.sleep(60)
        sleeps += 1
    else:
        sleeps = 0

IPython.start_ipython(['--quick', '--no-banner'], user_ns=globals())

