#!/usr/bin/python3

import re
import mmap
import pickle
import subprocess
from hashlib import blake2b
from operator import itemgetter
from datetime import datetime, timedelta, timezone
from contextlib import closing
from collections import Counter, defaultdict

import simhash
import IPython
from colorama import Fore
from openpyxl import load_workbook

from canvas import *
from xlinfo import filesinzips, workbook_props, xml_props
from xlsx2csv import process_cells, RowVisitor, SHEETSEP
from allgrades import fetch_grades
from wherelink import haslink, links_desc, file_haslink
from uniquecells import cleanval, CellCollector, filerpts, pairs_few, most_shared
from modderupdate import checkadd, Status, studict, fileinfo, writeout

USAGE = 'Arguments: [quizid(s)] [submission zip file(s)] [original Module file]'

sys.excepthook = IPython.core.ultratb.FormattedTB(mode='Verbose', color_scheme='Linux')

def numsonly(string):
    return ''.join(filter(str.isdigit, string))

def filetime(path):
    return datetime.fromtimestamp(Path(path).stat().st_mtime)

def basedir():
    return Path(__file__).parent.resolve()

def curdir():
    return Path.cwd()

def inbasedir():
    return curdir() == basedir()

def changetodir(dirname):
    os.makedirs(dirname, exist_ok=True)
    os.chdir(dirname)

def inemptydir():
    return not os.listdir()

def isgoodsubfile(subfile, exdate, modnum):
    subpat = re.compile(r"submissions \([0-9]+\)\.zip")
    nums = numsonly(subfile.stem)
    if not nums or subpat.fullmatch(subfile.name):
        return filetime(subfile).date() >= exdate
    return nums == modnum

def get_args(argv=sys.argv):
    quizids = []
    arg = 1
    while arg < len(argv) and argv[arg].isnumeric():
        quizids.append(int(argv[arg]))
        arg += 1
    if not quizids:
        exams = todays_exams()
        quizids = [e['quiz_id'] for e in exams]
    else:
        exams = dated_assigns('exams.json')
        exams = [ex for ex in exams if ex['quiz_id'] in quizids]
    if not exams:
        sys.exit('Could not find any quiz IDs. ' + USAGE)
    modnum = numsonly(exams[0]['name'])
    exdate = exams[0]['date']
    if argv[-1].endswith('.xlsx'):
        lastsub = len(argv) - 1
        origfile = Path(argv[-1]).resolve()
    else:
        lastsub = len(argv)
        origfiles = list((basedir() / 'original').glob(f'Module_{modnum}_*.xlsx'))
        if len(origfiles) == 1:
            origfile = origfiles[0]
        else:
            sys.exit('Please specify original module file')
    if arg < lastsub:
        subfiles = argv[arg:lastsub]
    else:
        globfiles = Path('~/Downloads').expanduser().glob('submissions*.zip')
        subfiles = [sf for sf in globfiles if isgoodsubfile(sf, exdate, modnum)]
        if not subfiles or not yesno(f'Using files {subfiles}. OK? '):
            sys.exit('Please specify downloaded submissions zip')
    print(f'Using quiz IDs {quizids}, submission zips {subfiles}, original file {origfile}', file=sys.stderr)
    return exams, subfiles, origfile, modnum

def blakesum(buf):
    bsum = blake2b(buf, digest_size=24)
    return bsum.hexdigest()

def bsum_fid(fid):
    return blakesum(mmap.mmap(fid.fileno(), 0, access=mmap.ACCESS_READ))

def bsum_mem(bytio):
    return blakesum(bytio.getbuffer())

def get_mime(filename):
    result = subprocess.run(['file', '-b', '--mime-type', filename], capture_output=True)
    return result.stdout.decode().rstrip()

def xls2xlsx(fdata):
    filename = fdata.name
    if os.path.exists(filename + 'x'):
        print(f'{filename}x found', file=sys.stderr)
    else:
        print(f'Converting {filename} to xlsx', file=sys.stderr)
        with open(filename, 'xb') as out:
            out.write(fdata.getbuffer())
        fdata.close()
        if get_mime(filename) == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            print("Appears to be misnamed .xlsx file", file=sys.stderr)
            os.rename(filename, filename + 'x')
        else:
            subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx', filename],
                    stdout=subprocess.DEVNULL)
    return open(filename + 'x', 'rb')

def filesize_mem(fdata):
    return fdata.getbuffer().nbytes

def filesize_fid(fid):
    return os.fstat(fid.fileno()).st_size

class BlakeHasher(RowVisitor):
    def __init__(self):
        self.bsum = blake2b(digest_size=24)
    def __call__(self, row):
        rowstr = ','.join(str(c) if c is not None else '' for c in row).rstrip(',') + '\n'
        self.bsum.update(rowstr.encode())
    def newsheet(self):
        self.bsum.update(SHEETSEP.encode())
    def value(self):
        return self.bsum.hexdigest()

class SimHasher(RowVisitor):
    def __init__(self):
        self.hashvector = []
    @staticmethod
    def hashval(shingle):
        return simhash.unsigned_hash(shingle.encode())
    def __call__(self, row):
        rowcells = [cleanval(c) for c in row if c is not None]
        if len(rowcells) >= 3:
            shingles = [' '.join(cells) for cells in simhash.shingle(rowcells, 3)]
        else:
            shingles = [' '.join(rowcells)]
        self.hashvector += [self.hashval(s) for s in shingles]
    def value(self):
        return simhash.compute(self.hashvector)

def process_workbook(xlhash, filename, workbook):
    cells, csvhash, simhash = process_cells(workbook, [CellCollector(), BlakeHasher(), SimHasher()])
    info = workbook_props(workbook, filename, xlhash, csvhash, simhash)
    return cells, info

def process_file(filename):
    with open(filename, 'rb') as fid:
        xlhash = bsum_fid(fid)
        if hasattr(filename, 'name'):
            filename = filename.name
        with closing(load_workbook(fid, read_only=True)) as workbook:
            return process_workbook(xlhash, filename, workbook)

def report_nonxlsx(filename):
    print('Not a xlsx file: ' + filename, file=sys.stderr)
    try:
        stuid = fileinfo(filename).stuid
    except ValueError: #filename not in Canvas name_sid_sub_etc format
        return # no known student to report
    reports[stuid].append(f'Non-xlsx file: {filename}')

def checktemp(filename, size):
    stuid = fileinfo(filename).stuid
    badname = '~' in filename
    small = size < 500
    if small and badname:
        reports[stuid].append('Temp file')
    elif badname:
        reports[stuid].append(f'Big temp file? {filename}')
    elif small:
        reports[stuid].append(f'Temp file? {filename}')
    else:
        reports[stuid].append(f'Not an Excel file? {filename}')

def pop_print_report(reports, stu):
    secname = stu.get('section','')
    stuid = stu['id']
    try:
        score = round(grades[stuid])
    except:
        score = ''
    namegrad = f"{stu.get('name','')} ({score})"
    print(f'{secname:4} {namegrad:26} ' + ', '.join(reports.pop(stuid)))

def dumb_lastname(tch):
    return tch['name'].split()[1]

def sorted_teachers():
    return sorted(teachers, key=dumb_lastname)

def print_reports(reports):
    studs = [studict[stuid] for stuid in reports]
    for tch in sorted_teachers():
        print(tch['name'])
        for sec in sorted([sec[1] for sec in tch['sections']]):
            secstuds = [stu for stu in studs if stu['section'] == sec]
            for stu in sorted(secstuds, key=itemgetter('sortable_name')):
                pop_print_report(reports, stu)
            if secstuds:
                print()
    if reports:
        print('Leftovers?!')
        for stuid in reports:
            pop_print_report(studict[stuid])

def reportmulti(stuid):
    msg = 'Multiple files?'
    if msg not in reports[stuid]:
        reports[stuid].append(msg)

def checkstuid(info, infos):
    codename, stuid = fileinfo(info.filename)
    stuids[stuid] += 1
    if stuids[stuid] > 1:
        print(f'{codename} seen {stuids[stuid]} times')
        if info.xlhash == originfo.xlhash:
            print('This one is unmodified, ignoring')
            return False
        prev = [inf for inf in infos if fileinfo(inf.filename).stuid == stuid]
        if prev[0].xlhash == originfo.xlhash: # Only the first added could be unmodified
            infos.remove(prev[0])
            print(f'Removing unmodified file {prev[0].filename}')
        elif any(p.xlhash == info.xlhash for p in prev):
            print('This one is identical to a previously seen file for this student; ignoring.')
            return False
        else:
            reportmulti(stuid)
    return True

def checksize(f):
    stuid = fileinfo(f.name).stuid
    size = filesize_mem(f)
    if size < 6000:
        reports[stuid].append(f'Small file, {size} bytes')

def checklink(f):
    stuid = fileinfo(f.name).stuid
    if file_haslink(f):
        wb = load_workbook(f, read_only=True)
        reports[stuid].append('Links to ' + links_desc(wb))
        wb.close()

def quickinfos(subfiles):
    infos = []
    for f in filesinzips(subfiles):
        xp = xml_props(f)
        if xp:
            xp.xlhash = bsum_mem(f)
            if checkstuid(xp, infos):
                infos.append(xp)
                checksize(f)
                checklink(f)
        else:
            checktemp(f.name, filesize_mem(f))
    return infos

def reportinfo(info):
    global originfo
    codename, stuid = fileinfo(info.filename)
    docreatmsg = False
    creatmsg = 'Created'
    if info.creator != originfo.creator:
        docreatmsg = True
        creatmsg += ' by ' + Fore.RED + info.creator + Fore.RESET
    if info.creation != originfo.creation:
        docreatmsg = True
        if abs(info.creation - originfo.creation) < timedelta(days=1):
            creatmsg += ' on ' + Fore.RED + f'{info.creation:%x %X}' + Fore.RESET
        else:
            creatmsg += ' on ' + Fore.RED + f'{info.creation:%x}' + Fore.RESET
    if docreatmsg:
        reports[stuid].append(creatmsg)

    domodmsg = False
    modmsg = 'Last modified'
    stat = checkadd(stuid, info.modder)
    # Status.Found, Status.Boo, Status.DNE, Status.Approved, Status.Unknown
    if stat is Status.DNE:
        return
    stu = studict[stuid]
    if stat is Status.Unknown:
        domodmsg = True
        modmsg += ' by ' + Fore.RED + info.modder + Fore.RESET
    elif stat is Status.Boo and info.xlhash == originfo.xlhash:
        reports[stuid].append('Unmodified')
    elif stat is Status.Boo and info.modified == originfo.modified:
        reports[stuid].append('Metadata says unmodified')
    elif stat is Status.Boo:
        reports[stuid].append(f'Unmodified wrong spreadsheet? Last modified by {info.modder} on {info.modified:%x %X}')
    if info.modified > examtime + timedelta(days=1) or info.modified < examtime - timedelta(days=9) and stat is not Status.Boo:
        if not domodmsg:
            modmsg += ' by ' + info.modder
        modmsg += ' on ' + Fore.RED + f'{info.modified:%x}' + Fore.RESET
        domodmsg = True
    if domodmsg:
        reports[stuid].append(modmsg)

def reportidentical(hasht):
    hashes = Counter(i.asdict()[hasht] for i in infos)
    for shash, count in hashes.most_common():
        if count <= 1:
            break
        print(f'Identical {hasht}: ' + ', '.join(i.filename for i in infos if i.asdict()[hasht] == shash))

# TODO: download the submissions here
# Note: assignment json has a submissions_download_url which is purported to let you download the zip of all submissions
# However, it only gives an HTML page with authentication error :(
# https://community.canvaslms.com/thread/10824

# Future plan: make working directory /tmp/examcheck; remove any files
# download each submission as it comes in, a la secdownload
# process the file in memory. instead of writing csv, save the simhash, xlhash, csvhash
# if there's any problem to report, write out the .xlsx and .csv files.
# I guess near-dups are found at the end, so re-download cluster members after comparing simhashes?
# also re-download uniquecell cluster members.

if __name__ == '__main__':
    exams, subfiles, origfile, modnum = get_args()
    report = None

    if inbasedir():
        mdir = 'mod' + numsonly(exams[0]['name'])
        report = Path(mdir + '-report')
        report.touch()
        changetodir(mdir)
        print('Using directory ' + mdir, file=sys.stderr)
        report = Path(os.pardir, report)

    if not inemptydir():
        if not yesno(f'Current directory {curdir()} is not empty. Proceed (may clobber files)? '):
            sys.exit('Aborted.')

    Path('orig.xlsx').symlink_to(origfile)
    if len(subfiles) == 1:
        Path('subs.zip').symlink_to(subfiles[0])
    if report:
        Path('report').symlink_to(report)

    teachers = load_file('teachers.json', json.load)
    origcells, originfo = process_file(origfile)

    cellfiles = defaultdict(list) # map cell_content : files
    grades = fetch_grades([ex['quiz_id'] for ex in exams]) # map student_id : score
    stuids = Counter()
    xlhashes = {}
    reports = defaultdict(list) # map stuid : strings

    examtime = datetime.combine(exams[0]['date'], Time(17)).astimezone(timezone.utc) # 5 pm
    if len({ex['date'] for ex in exams}) > 1:
        print(f'Warning: exams on different dates. Using first exam {exams[0]["date"]}', file=sys.stderr)

    warnings.filterwarnings('ignore', '.*invalid specification.*', UserWarning, 'openpyxl')
    warnings.filterwarnings('ignore', 'Unknown extension is not supported.*', UserWarning, 'openpyxl')

    infos = quickinfos(subfiles)
    for info in infos:
        reportinfo(info)

    print_reports(reports.copy())
    print('---------------------')

    for file in filesinzips(subfiles):
        try:
            info = next(i for i in infos if i.filename in (file.name, file.name + 'x'))
            xlhash = info.xlhash
        except StopIteration:
            print(f'Could not find quickinfo for file {file.name}', file=sys.stderr)
            info = None
            xlhash = bsum_mem(file)
        codename, stuid = fileinfo(file.name)
        size = filesize_mem(file)
        if xlhash in xlhashes:
            prev = xlhashes[xlhash]
            print(f'File {file.name} identical to previously seen file {prev.filename}')
            continue
        if file.name.endswith('.xls'):
            file = xls2xlsx(file)
        elif not file.name.endswith('.xlsx'):
            report_nonxlsx(file.name)
            continue
        try:
            wb = load_workbook(file, read_only=True)
        except Exception as e:
            print(file.name, 'is not an xlsx file?', e, file=sys.stderr)
            checktemp(file.name, size)
            continue
        thecells, theinfo = process_workbook(xlhash, file.name, wb)
        for cval in thecells - origcells:
            cellfiles[cval].append(file.name)
        if info and info != theinfo.replace(csvhash='', simhash=0):
            print('different Info:', info, theinfo, file=sys.stderr)
        if info:
            info.csvhash = theinfo.csvhash
            info.simhash = theinfo.simhash
        else:
            infos.append(theinfo)
        xlhashes[xlhash] = theinfo
        if haslink(wb):
            reports[stuid].append('Links to ' + links_desc(wb))
        wb.close()
        file.close()

    pdata = {'origcells': origcells,
            'originfo': originfo,
            'cellfiles': cellfiles,
            'xlhashes': xlhashes,
            'infos': infos,
            'reports': reports,
            'grades': grades}
    with open(f'modd{modnum}.pkl', 'wb') as fid:
        pickle.dump(pdata, fid, protocol=pickle.HIGHEST_PROTOCOL)

    writeout()
    pairs = simhash.find_all([i.simhash for i in infos], 6, 4) # blocks >= maxdist + 1; maxdist 1 to 64
    print_reports(reports.copy())

    reportidentical('xlhash')
    reportidentical('csvhash')
    reportidentical('simhash')

    rpts = filerpts(cellfiles)
    most_shared(rpts, cellfiles)
    pairs_few(rpts, cellfiles)

    IPython.start_ipython(['--quick', '--no-banner'], user_ns=globals())

# Make command interpreter (import cmd)
# <stu> is codename with tab-completion, or stuid
# report <stu> [message]
# get <stu> -- downloads and converts to CSV
# open <stu> -- gets if necessary, opens in libreoffice
# diff <stu1> <stu2> -- opens icdiff of CSVs in a pager
