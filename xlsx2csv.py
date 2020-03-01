#!/usr/bin/python3

import os
import csv
import sys
from openpyxl import load_workbook

SHEETSEP = '----------\n'
csv.register_dialect('newline', lineterminator='\n')

def process_cells(workbook, rowvisitors):
    """Visit all rows of the workbook, passing to given RowVisitors,
    and return their values."""
    for ws in workbook.worksheets:
        ws.reset_dimensions()
        maxrow = max(n for n,row in enumerate(ws.values) if any(row))
        for row in ws.iter_rows(max_row=maxrow+1, values_only=True):
            for visit in rowvisitors:
                visit(row)
        for visit in rowvisitors:
            visit.newsheet()
    return [rv.value() for rv in rowvisitors]

class RowVisitor:
    def newsheet(self):
        pass

class CSVPrinter(RowVisitor):
    def __init__(self, filename):
        self.fid = open(filename, 'w', newline='')
        self.writer = csv.writer(self.fid, 'newline')
    def __call__(self, row):
        maxcol = max((n for n,c in enumerate(row) if c is not None), default=-1)
        self.writer.writerow(row[:maxcol+1])
    def newsheet(self):
        self.fid.write(SHEETSEP)
    def value(self):
        self.fid.close()

def get_args(argv=sys.argv):
    if len(argv) < 2:
        sys.exit('At least one argument, infile [outfile]')
    infile = argv[1]
    if len(argv) == 3:
        outfile = argv[2]
    elif len(argv) > 3:
        sys.exit('Max two arguments, infile [outfile]')
    else:
        root, ext = os.path.splitext(infile)
        outfile = root + '.csv'
    return infile, outfile

if __name__ = '__main__':
    infile, outfile = get_args()
    wb = load_workbook(filename=infile, read_only=True)
    process_cells(wb, [CSVPrinter(outfile)])
